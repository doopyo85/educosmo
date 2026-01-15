import openpyxl
import os

excel_path = r"c:\Users\User\Documents\pioneer\educosmo\s3\ContentMap.xlsx"

if not os.path.exists(excel_path):
    print(f"Error: File not found at {excel_path}")
    exit(1)

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Print first row (headers)
headers = [cell.value for cell in ws[1]]
print("Headers:", headers)

# Print first few rows to see data sample
print("\nFirst 3 rows:")
for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
    print(row)
